"""Print the exact characters of each sheet name in an Excel file.

Usage: python scripts/inspect_sheet_names.py [path/to/file.xlsx]
If no path is given, opens a file dialog.
"""

import sys
import unicodedata

import openpyxl

EXPECTED = ["Tickets ADUS Tickets crea... 1", "Tickets ADUS Tickets crea... 2"]


def describe(name: str) -> None:
    print(f"  repr: {name!r}")
    for ch in name:
        marker = "" if ord(ch) < 128 else "   <-- NON-ASCII"
        print(f"    U+{ord(ch):04X} {ch!r:6} {unicodedata.name(ch, '?')}{marker}")


def main() -> None:
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        from tkinter import filedialog

        path = filedialog.askopenfilename(
            title="Select Excel file to inspect",
            filetypes=[("Excel files", "*.xlsx *.xlsm")],
        )
        if not path:
            print("No file selected.")
            return

    wb = openpyxl.load_workbook(path, read_only=True)
    print(f"File: {path}")
    print(f"Sheets found: {len(wb.sheetnames)}\n")
    for name in wb.sheetnames:
        print(f"Sheet: {name}")
        describe(name)
        matches = [e for e in EXPECTED if e == name]
        if matches:
            print("  ✓ exactly matches a config sheet name")
        else:
            close = [e for e in EXPECTED if e.replace(" ", "") == name.replace(" ", "")]
            if close:
                print("  ⚠ matches config except for whitespace differences!")
            else:
                print("  ✗ does not match any config sheet name")
        print()


if __name__ == "__main__":
    main()
